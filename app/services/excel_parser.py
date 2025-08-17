from datetime import datetime
import openpyxl
import json

from flask_login import current_user

from app.models import Interface, InterfaceParam, TestCase, db, ParamType  # 确保导入正确的模型路径
import logging
from sqlalchemy.exc import IntegrityError

# 初始化日志
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def parse_excel(file, project_id, creator_id):
    """
    解析 Excel 接口文档，关联接口参数与测试用例，支持区分参数类型（query/body/header 等）
    :param file: 上传的文件对象（Excel）
    :param project_id: 项目 ID（必须非空，关联 Project 表）
    :param creator_id: 测试用例创建者 ID（关联 User 表）
    :return: 解析结果字典，包含 success、parse_success、error 等键
    """
    # 前置校验：确保 project_id 非空
    if not project_id:
        return {
            "success": False,
            "parse_success": False,
            "error": "project_id 不能为空！解析终止"
        }

    try:
        # 加载 Excel 文件（带数据类型转换）
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        logger.debug(f"加载 Excel 成功，工作表: {sheet.title}，总行数: {sheet.max_row}")
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # 定义需要校验的表头展示名称以及对应的匹配关键字
        required_header_info = {
            "接口名称": ["接口名称", "api名称"],
            "URL": ["url", "地址"],
            "method": ["method", "方法", "请求方式"],
            "请求头": ["请求头", "headers"],
            "请求体": ["请求体", "body"],
            "响应结果": ["响应结果", "response"]
        }

        # 检查表头是否包含所有必需字段
        missing_headers = []
        for display_name, keywords in required_header_info.items():
            has_header = any(
                kw in str(header).lower().strip()
                for header in header_row
                for kw in keywords
            )
            if not has_header:
                missing_headers.append(display_name)
        if missing_headers:
            return {
                "success": False,
                "parse_success": False,
                "error": f"Excel 缺少必需表头：{', '.join(missing_headers)}"
            }

        # 读取表头并建立映射（支持多语言/别名兼容）
        header_mapping = {}
        for idx, header in enumerate(header_row):
            if not header:
                continue
            header_lower = header.lower().strip()
            if any(key in header_lower for key in ["接口名称", "api名称"]):
                header_mapping['interface_name'] = idx
            elif any(key in header_lower for key in ["url", "地址"]):
                header_mapping['url'] = idx
            elif any(key in header_lower for key in ["method", "方法", "请求方式"]):
                header_mapping['method'] = idx
            elif any(key in header_lower for key in ["请求头", "headers"]):
                header_mapping['request_header'] = idx
            elif any(key in header_lower for key in ["请求体", "body"]):
                header_mapping['request_body'] = idx
            elif any(key in header_lower for key in ["响应结果", "response"]):
                header_mapping['response_result'] = idx
            elif any(key in header_lower for key in ["参数类型", "param type"]):
                header_mapping['param_type'] = idx

        interface_count = 0  # 统计新增接口数
        param_count = 0  # 统计参数总数
        case_count = 0  # 统计测试用例数

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = list(row)

            # 提取核心字段（空值校验）
            interface_name = row_data[header_mapping.get('interface_name')]
            url = row_data[header_mapping.get('url')]
            method = row_data[header_mapping.get('method')]
            if not (interface_name and url and method):
                logger.warning(f"第{row_num}行数据缺失核心字段，跳过")
                continue
            method = method.upper().strip()  # 统一转为大写

            # 解析 param_type
            param_type_str = row_data[header_mapping.get('param_type')] or 'body'
            param_type = None
            try:
                param_type = ParamType(param_type_str.lower().strip())
            except ValueError:
                logger.warning(f"第{row_num}行 param_type 无效：{param_type_str}，已使用默认值 'body'")
                param_type = ParamType('body')

            # 关联或创建接口（严格关联 project_id）
            interface = Interface.query.filter_by(
                project_id=project_id,
                url=url,
                method=method
            ).first()

            if not interface:
                # 处理请求头
                request_header_raw = row_data[header_mapping.get('request_header')] or '{}'
                try:
                    request_header = json.loads(request_header_raw)
                except json.JSONDecodeError:
                    request_header = request_header_raw
                    logger.warning(f"第{row_num}行请求头非 JSON 格式，已保留原始值: {request_header_raw}")

                # 创建新接口
                interface = Interface(
                    project_id=project_id,
                    interface_name=interface_name,
                    url=url,
                    method=method,
                    request_header=json.dumps(request_header) if isinstance(request_header, (dict, list)) else request_header,
                    create_time=datetime.now()
                )
                db.session.add(interface)
                db.session.commit()
                interface_count += 1
                logger.debug(f"新增接口：ID={interface.interface_id}, 名称={interface_name}")

                # 解析请求体参数
                request_body_raw = row_data[header_mapping.get('request_body')] or '{}'
                try:
                    request_body = json.loads(request_body_raw)
                except json.JSONDecodeError:
                    request_body = request_body_raw
                    logger.warning(f"第{row_num}行请求体非 JSON 格式，已保留原始值: {request_body_raw}")

                # 递归解析嵌套参数
                nested_count = parse_nested_params(
                    interface_id=interface.interface_id,
                    data=request_body if isinstance(request_body, dict) else {},
                    param_type=param_type
                )
                param_count += nested_count

            # 处理响应结果
            response_result_raw = row_data[header_mapping.get('response_result')] or '{}'
            try:
                response_result = json.loads(response_result_raw)
            except json.JSONDecodeError:
                response_result = response_result_raw
                logger.warning(f"第{row_num}行响应结果非 JSON 格式，已保留原始值: {response_result_raw}")

            # 创建测试用例
            case = create_test_case(
                interface=interface,
                request_body_str=json.dumps(request_body) if isinstance(request_body, (dict, list)) else request_body,
                response_result_str=json.dumps(response_result) if isinstance(response_result, (dict, list)) else response_result,
                creator_id=creator_id
            )
            if case:
                case_count += 1
                db.session.add(case)

        # 提交所有变更
        db.session.commit()
        return {
            "success": True,
            "parse_success": True,
            "message": f"Excel 解析完成！新增接口 {interface_count} 个，参数 {param_count} 个，测试用例 {case_count} 个"
        }

    except IntegrityError as e:
        db.session.rollback()
        return {
            "success": False,
            "parse_success": False,
            "error": f"数据库冲突: {str(e)}"
        }
    except Exception as e:
        db.session.rollback()
        return {
            "success": False,
            "parse_success": False,
            "error": f"Excel 解析失败: {str(e)}"
        }


def parse_nested_params(interface_id, data, param_type, parent_key=''):
    """
    递归解析嵌套参数，支持 JSON 层级，转换数据类型名称
    :param interface_id: 接口 ID
    :param data: JSON 数据（dict）
    :param param_type: 参数类型（ParamType 枚举）
    :param parent_key: 嵌套层级
    :return: 解析的参数数量
    """
    def normalize_data_type(data_type):
        type_mapping = {
            'str': 'string',
            'int': 'number',
            'float': 'number',
            'dict': 'object',
            'list': 'array',
            'bool': 'boolean',
            'NoneType': 'null'
        }
        return type_mapping.get(data_type.lower(), data_type)

    count = 0
    if not isinstance(data, dict):
        return count

    for key, value in data.items():
        full_key = f"{parent_key}.{key}" if parent_key else key
        raw_data_type = type(value).__name__
        data_type = normalize_data_type(raw_data_type)

        param = InterfaceParam(
            interface_id=interface_id,
            param_name=full_key,
            param_type=param_type.value,
            data_type=data_type,
            is_required=True,
            parent_key=parent_key,
            example_value=str(value)
        )
        db.session.add(param)
        count += 1

        if isinstance(value, dict):
            count += parse_nested_params(interface_id, value, param_type, full_key)
    return count

#
# def create_test_case(interface, request_body_str, response_result_str, creator_id):
#     """
#     创建测试用例，关联接口参数
#     :param interface: Interface 实例
#     :param request_body_str: 请求体内容
#     :param response_result_str: 响应结果内容
#     :param creator_id: 用例创建者 ID
#     :return: TestCase 实例或 None
#     """
#     try:
#         request_body = json.loads(request_body_str) if request_body_str else {}
#         response_result = json.loads(response_result_str) if response_result_str else {}
#
#         param_mapping = {}
#         for param in interface.params:
#             keys = param.param_name.split('.')
#             current = request_body
#             try:
#                 for k in keys:
#                     current = current[k]
#                 param_mapping[param.param_name] = current
#             except (KeyError, TypeError):
#                 continue
#
#         test_case = TestCase(
#             interface_id=interface.interface_id,
#             case_name=f"用例_{interface.interface_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}",
#             param_values=json.dumps(param_mapping),
#             expected_result=json.dumps(response_result),
#             assert_rule=generate_assert_rule(response_result),
#             creator_id=creator_id,
#             create_time=datetime.now()
#         )
#         return test_case
#     except json.JSONDecodeError as e:
#         logger.error(f"创建测试用例失败: {e}")
#         return None
#
#
# def generate_assert_rule(response_result):
#     """自动生成简单断言规则"""
#     if 'status' in response_result:
#         return f"response['status'] == '{response_result['status']}'"
#     if 'code' in response_result:
#         return f"response['code'] == {response_result['code']}"
#     return "response.get('status') == 'success'"