import os
import logging
from datetime import datetime
import openpyxl
from app.models import Interface, InterfaceParam, db, Project, User
from openpyxl.utils import column_index_from_string
import json

# 配置日志，方便调试
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def parse_excel(file, project_id):
    """
    解析Excel接口文档并存储到数据库（带调试输出）
    :param file: 上传的文件对象（Flask request.files['file']）
    :param project_id: 项目ID
    :return: 解析结果字符串
    """
    try:
        # 1. 保存上传的文件到临时目录（方便调试时查看文件内容）
        temp_dir = "temp_uploads"
        os.makedirs(temp_dir, exist_ok=True)  # 创建临时目录

        # 生成唯一文件名（避免覆盖）
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = os.path.join(temp_dir, filename)
        file.save(file_path)
        logger.debug(f"文件已保存到临时路径: {file_path}")

        # 2. 加载Excel文件
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        logger.debug(f"加载Excel成功，工作表名称: {sheet.title}，总行数: {sheet.max_row}")

        # 3. 表头映射（可根据实际Excel调整）
        header_mapping = {
            'interface_name': 0,   # 接口名称（A列）
            'url': 1,              # URL（B列）
            'method': 2,           # 请求方法（C列）
            'request_header': 3,   # 请求头（D列）
            'request_body': 4,     # 请求体（E列）
            'response_result': 6   # 响应结果（G列）
        }

        # 4. 校验表头
        if sheet.max_column < len(header_mapping):
            error_msg = f"Excel格式错误：至少需要{len(header_mapping)}列，实际只有{sheet.max_column}列"
            logger.error(error_msg)
            return error_msg

        # 5. 解析数据并存储
        interface_list = []  # 用于记录导入的接口（调试用）
        param_list = []      # 用于记录导入的参数（调试用）

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = list(row)
            # 填充空值，避免索引错误
            row_data += [None] * (len(header_mapping) - len(row_data))

            # 提取接口信息
            interface_name = row_data[header_mapping['interface_name']]
            url = row_data[header_mapping['url']]
            method = row_data[header_mapping['method']]

            if not all([interface_name, url, method]):
                logger.warning(f"第{row_num}行数据不完整，跳过解析")
                continue

            # 创建接口记录
            interface = Interface(
                project_id=project_id,
                interface_name=interface_name,
                url=url,
                method=method,
                request_header=row_data[header_mapping['request_header']]
            )
            db.session.add(interface)
            db.session.flush()  # 暂存获取ID
            interface_list.append({
                "interface_id": interface.interface_id,
                "name": interface_name,
                "url": url,
                "method": method,
                "row": row_num
            })

            # 解析请求体JSON获取参数信息
            request_body_str = row_data[header_mapping['request_body']]
            if request_body_str:
                try:
                    request_body = json.loads(request_body_str)
                    for param_name, param_value in request_body.items():
                        param_type = 'body'  # 这里简单假设请求体参数类型为body，可根据实际调整
                        data_type = type(param_value).__name__  # 简单获取数据类型，可进一步优化
                        param = InterfaceParam(
                            interface_id=interface.interface_id,
                            param_name=param_name,
                            param_type=param_type,
                            data_type=data_type,
                            is_required=True,  # 这里简单假设为必填，可根据实际添加逻辑判断
                            example_value=str(param_value)
                        )
                        db.session.add(param)
                        param_list.append({
                            "param_id": param.param_id,
                            "name": param_name,
                            "interface_id": interface.interface_id,
                            "row": row_num
                        })
                except json.JSONDecodeError as e:
                    logger.error(f"第{row_num}行请求体JSON解析错误: {e}")

        # 提交事务
        db.session.commit()
        logger.debug("数据已成功提交到数据库")

        # 6. 输出调试信息（返回给前端或打印到控制台）
        debug_info = (
            f"解析完成！\n"
            f"临时文件路径: {file_path}\n"
            f"新增接口数: {len(interface_list)}\n"
            f"新增参数数: {len(param_list)}\n"
            f"接口详情: {interface_list}\n"
            f"参数详情: {param_list}"
        )
        logger.info(debug_info)
        return debug_info  # 返回详细信息供调试

    except Exception as e:
        db.session.rollback()  # 出错时回滚事务
        error_msg = f"解析失败: {str(e)}"
        logger.error(error_msg, exc_info=True)  # 记录详细错误堆栈
        return error_msg